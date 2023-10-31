import pandas as pd
import numpy as np
from openpyxl.utils import coordinate_to_tuple, get_column_letter


def get_cell_contents_row(cur_cell):
    cell_features = {
        "coordinate": cur_cell.coordinate,
        "is_empty": cur_cell.value is None,
        "is_string": cur_cell.data_type in ['s', 'str'],
        "is_merged": type(cur_cell).__name__ == 'MergedCell',
        "is_bold": cur_cell.font.b,
        "is_italic": cur_cell.font.i,
        "left_border": cur_cell.border.left is not None,
        "right_border": cur_cell.border.right is not None,
        "top_border": cur_cell.border.top is not None,
        "bottom_border": cur_cell.border.bottom is not None,
        "is_filled": cur_cell.fill.patternType is not None,
        "horizontal_alignment": cur_cell.alignment.horizontal is not None,
        "left_horizontal_alignment": cur_cell.alignment.horizontal == 'left',
        "right_horizontal_alignment": cur_cell.alignment.horizontal == 'right',
        "center_horizontal_alignment": cur_cell.alignment.horizontal == 'center',
        "wrapped_text": cur_cell.alignment.wrapText,
        "indent": cur_cell.alignment.indent != 0,
        "formula": cur_cell.data_type == 'f'
    }

    return cell_features


def get_cell_features(ws) -> pd.DataFrame:
    """
    На вход: адрес Excel файла (опционально адрес файла с разметкой зон)
    На выход: таблица с характеристиками каждой непустой ячейки файла
    Зачем: анализ hard values, умножения на ноль, использованных формул и т.д.
    """

    data = []

    for row in ws.iter_rows():
        for cell in row:
            data.append(get_cell_contents_row(cell))

    result_df = pd.DataFrame(data)

    return result_df


def apply_neighbouring_features(df, range_of_neighbors):
    """
    range_of_neighbors: tuple(height, width)
    """
    working_df = df.copy()
    working_df = working_df.set_index("coordinate")

    neighbouring_features = ["is_empty", "is_merged", "is_bold", "is_italic",
                             "left_border", "right_border", "top_border", "bottom_border",
                             "is_filled", "wrapped_text", "indent", "formula"]

    h, w = range_of_neighbors
    for index, row in working_df.iterrows():
        row = row.copy()
        r, c = coordinate_to_tuple(index)
        for i in range(-h, h+1):
            for j in range(-w, w+1):
                if i == 0 and j == 0:
                    continue
                new_r = r + i
                new_c = c + j
                if new_r > 0 and new_c > 0:
                    new_coordinate = get_column_letter(new_c) + str(new_r)
                    if new_coordinate in working_df.index:
                        for neighbouring_feature in neighbouring_features:
                            new_column_name = neighbouring_feature + str(i) + "_" + str(j)
                            neighbouring_value = working_df.loc[new_coordinate, neighbouring_feature]
                            if new_column_name not in working_df.columns:
                                working_df[new_column_name] = 0
                                working_df[new_column_name] = working_df[new_column_name].astype(int)
                            working_df.loc[index, new_column_name] = neighbouring_value

    working_df = working_df.reset_index()
    return working_df

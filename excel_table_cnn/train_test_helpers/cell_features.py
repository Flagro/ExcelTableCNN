import pandas as pd
import xlrd
import openpyxl
from openpyxl.utils.cell import get_column_letter


def get_cell_features_xls(cur_cell, row_idx, col_idx, sheet, book):
    xf = book.xf_list[cur_cell.xf_index]
    font = book.font_list[xf.font_index]
    
    cell_features = {
        "coordinate": get_column_letter(col_idx) + str(row_idx),
        "is_empty": cur_cell.ctype == xlrd.XL_CELL_EMPTY,
        "is_string": cur_cell.ctype == xlrd.XL_CELL_TEXT,
        "is_merged": (cur_cell.row, cur_cell.col) in sheet.merged_cells,
        "is_bold": font.bold,
        "is_italic": font.italic,
        "formula": cur_cell.ctype == xlrd.XL_CELL_FORMULA,
        # Some features might not have a direct equivalent in xlrd, so they are approximated or left out
        "left_border": False,
        "right_border": False,
        "top_border": False,
        "bottom_border": False,
        "is_filled": False,
        "horizontal_alignment": False,
        "left_horizontal_alignment": False,
        "right_horizontal_alignment": False,
        "center_horizontal_alignment": False,
        "wrapped_text": False,
        "indent": False
    }
    return cell_features


def get_cell_features_xlsx(cur_cell):
    cell_features = {
        "coordinate": (cur_cell.row, cur_cell.column),
        "is_empty": cur_cell.value is None,
        "is_string": cur_cell.data_type in ['s', 'str'],
        "is_merged": type(cur_cell).__name__ == 'MergedCell',
        "is_bold": cur_cell.font.b,
        "is_italic": cur_cell.font.i,
        "left_border": cur_cell.border.left is not None,
        "right_border": cur_cell.border.right is not None,
        "top_border": cur_cell.border.top is not None,
        "bottom_border": cur_cell.border.bottom is not None,
        "is_filled": cur_cell.fill.patternType is not None,
        "horizontal_alignment": cur_cell.alignment.horizontal is not None,
        "left_horizontal_alignment": cur_cell.alignment.horizontal == 'left',
        "right_horizontal_alignment": cur_cell.alignment.horizontal == 'right',
        "center_horizontal_alignment": cur_cell.alignment.horizontal == 'center',
        "wrapped_text": cur_cell.alignment.wrapText,
        "indent": cur_cell.alignment.indent != 0,
        "formula": cur_cell.data_type == 'f'
    }
    return cell_features


def get_table_features(file_path, sheet_name) -> pd.DataFrame:
    if file_path.endswith('.xls'):
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        ws = wb.sheet_by_name(sheet_name)
        data = []
        for row in range(ws.nrows):
            for col in range(ws.ncols):
                cell = ws.cell(row, col)
                data.append(get_cell_features_xls(cell, row, col, ws, wb))
    
    elif file_path.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows():
            for cell in row:
                data.append(get_cell_features_xlsx(cell))
    else:
        raise ValueError(f"Unsupported file format for file: {file_path}")

    result_df = pd.DataFrame(data)
    return result_df

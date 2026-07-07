"""Cell featurization: turn an openpyxl worksheet into an (H, W, C) float array.

v2 scheme, 30 channels per cell (``FEATURE_NAMES`` gives the order): the
17 original binary channels plus value-string, data-format and cell-format
groups (string statistics, number-format template classification, merge
direction, non-default colors). Non-binary channels are normalized to [0, 1].

The array always starts at A1 so that annotation coordinates remain valid
without offset bookkeeping; trailing all-default rows/columns are trimmed
(plus a small margin), and the used range is capped to protect against
sheets with huge stray formatting.
"""

import logging
import math
import re
from typing import List, Optional, Tuple

import numpy as np
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Bump when FEATURE_NAMES or their semantics change: invalidates cached tensors.
FEATURES_VERSION = "v2"

FEATURE_NAMES: List[str] = [
    # --- v1 binary channels (order preserved) ---
    "is_empty",
    "is_string",
    "is_merged",
    "is_bold",
    "is_italic",
    "left_border",
    "right_border",
    "top_border",
    "bottom_border",
    "is_filled",
    "horizontal_alignment",
    "left_horizontal_alignment",
    "right_horizontal_alignment",
    "center_horizontal_alignment",
    "wrapped_text",
    "indent",
    "formula",
    # --- v2: value-string statistics (normalized) ---
    "string_length",
    "digit_ratio",
    "letter_ratio",
    "has_percent_symbol",
    "has_decimal_point",
    # --- v2: number-format template classification ---
    "numeric_format",
    "date_format",
    "time_format",
    "format_length",
    # --- v2: cell-format extensions ---
    "merged_horizontal",
    "merged_vertical",
    "non_default_fill_color",
    "non_default_font_color",
]

NUM_FEATURES = len(FEATURE_NAMES)

_IS_EMPTY_IDX = FEATURE_NAMES.index("is_empty")

# Hard caps on the featurized used range. Real tables larger than this exist
# but are rare in the training corpus; stray formatting thousands of rows below
# the data is common and would otherwise explode memory.
DEFAULT_MAX_ROWS = 2048
DEFAULT_MAX_COLS = 512

# Rows/columns kept beyond the last non-default cell, so tables whose
# annotation extends slightly into empty cells aren't clipped.
TRIM_MARGIN = 2

_STRING_LENGTH_CAP = 1024.0
_FORMAT_LENGTH_CAP = 64.0


def default_cell_vector() -> np.ndarray:
    """Feature vector of an untouched cell: empty, everything else off."""
    vec = np.zeros(NUM_FEATURES, dtype=np.float32)
    vec[_IS_EMPTY_IDX] = 1.0
    return vec


def value_text(value) -> str:
    """Canonical text of a cell value, identical across backends.

    xlrd reports every number as float (30 -> 30.0) while openpyxl keeps
    ints; canonicalize so string statistics agree.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def text_stats(text: str) -> List[float]:
    """string_length, digit_ratio, letter_ratio, has %, has '.' — all [0,1]."""
    if not text:
        return [0.0, 0.0, 0.0, 0.0, 0.0]
    n = len(text)
    digits = sum(ch.isdigit() for ch in text)
    letters = sum(ch.isalpha() for ch in text)
    return [
        min(math.log1p(n) / math.log1p(_STRING_LENGTH_CAP), 1.0),
        digits / n,
        letters / n,
        float("%" in text),
        float("." in text),
    ]


def format_features(number_format: Optional[str]) -> List[float]:
    """numeric_format, date_format, time_format, format_length — from the
    cell's number-format template string (e.g. ``#,##0.00`` or ``yyyy-mm-dd``).

    Color sections (``[Red]``) and quoted literals are stripped before token
    matching so e.g. the ``d`` in ``[Red]`` doesn't read as a date token.
    """
    if not number_format or number_format.lower() == "general":
        return [0.0, 0.0, 0.0, 0.0]
    fmt = re.sub(r"\[[^\]]*\]", "", number_format.lower())
    fmt = re.sub(r'"[^"]*"', "", fmt)
    is_time = any(token in fmt for token in ("h", "s", "am/pm"))
    is_date = "y" in fmt or "d" in fmt or ("m" in fmt and not is_time)
    is_numeric = "0" in fmt or "#" in fmt
    return [
        float(is_numeric),
        float(is_date),
        float(is_time),
        min(len(number_format) / _FORMAT_LENGTH_CAP, 1.0),
    ]


def _openpyxl_color_is_nondefault(color) -> bool:
    # Dispatch on Color.type: accessing a mismatched attribute (e.g. .rgb on
    # a theme color) returns the Typed-descriptor's error *string*, not None.
    if color is None:
        return False
    color_type = getattr(color, "type", None)
    if color_type == "rgb":
        return color.rgb.upper() not in ("00000000", "FF000000", "FFFFFFFF")
    if color_type == "indexed":
        return color.indexed not in (0, 1, 64, 65)  # black/white/auto
    if color_type == "theme":
        return color.theme not in (0, 1)  # background-1 / text-1
    return False  # auto or unknown


def cell_feature_vector(cell, merged_extent: Optional[Tuple[bool, bool]]) -> List[float]:
    """Features of a single openpyxl cell, ordered as FEATURE_NAMES.

    ``merged_extent`` is None for unmerged cells, else (spans_multiple_cols,
    spans_multiple_rows) of the merged range the cell belongs to.
    """
    alignment = cell.alignment
    border = cell.border
    text = value_text(cell.value)
    merged_h = bool(merged_extent and merged_extent[0])
    merged_v = bool(merged_extent and merged_extent[1])
    return [
        float(cell.value is None),
        float(cell.data_type in ("s", "str")),
        float(merged_extent is not None),
        float(bool(cell.font.b)),
        float(bool(cell.font.i)),
        # ``border.left`` is always a Side object in openpyxl; only its
        # ``style`` says whether a border is actually drawn.
        float(border.left.style is not None),
        float(border.right.style is not None),
        float(border.top.style is not None),
        float(border.bottom.style is not None),
        float(cell.fill.patternType is not None),
        float(alignment.horizontal is not None),
        float(alignment.horizontal == "left"),
        float(alignment.horizontal == "right"),
        float(alignment.horizontal == "center"),
        float(bool(alignment.wrapText)),
        float(bool(alignment.indent)),
        float(cell.data_type == "f"),
        *text_stats(text),
        *format_features(cell.number_format),
        float(merged_h),
        float(merged_v),
        float(
            cell.fill.patternType is not None
            and _openpyxl_color_is_nondefault(cell.fill.fgColor)
        ),
        float(_openpyxl_color_is_nondefault(cell.font.color)),
    ]


def featurize_sheet(
    ws: Worksheet,
    max_rows: int = DEFAULT_MAX_ROWS,
    max_cols: int = DEFAULT_MAX_COLS,
) -> np.ndarray:
    """Featurize a worksheet into an (H, W, NUM_FEATURES) float32 array."""
    used_rows = ws.max_row or 1
    used_cols = ws.max_column or 1
    if used_rows > max_rows or used_cols > max_cols:
        logger.warning(
            "Sheet %r used range %dx%d exceeds cap %dx%d; clipping",
            ws.title, used_rows, used_cols, max_rows, max_cols,
        )
    n_rows = min(used_rows, max_rows)
    n_cols = min(used_cols, max_cols)

    arr = np.tile(default_cell_vector(), (n_rows, n_cols, 1))

    merged_extents = {}
    for merged_range in ws.merged_cells.ranges:
        extent = (
            merged_range.max_col > merged_range.min_col,
            merged_range.max_row > merged_range.min_row,
        )
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                merged_extents[(r, c)] = extent

    for row in ws.iter_rows(min_row=1, max_row=n_rows, min_col=1, max_col=n_cols):
        for cell in row:
            arr[cell.row - 1, cell.column - 1] = cell_feature_vector(
                cell, merged_extents.get((cell.row, cell.column))
            )

    return trim_trailing_default(arr)


def trim_trailing_default(arr: np.ndarray) -> np.ndarray:
    """Trim trailing rows/columns whose cells all equal the default vector."""
    non_default = (arr != default_cell_vector()).any(axis=2)
    if not non_default.any():
        return arr[:1, :1]
    last_row = int(np.nonzero(non_default.any(axis=1))[0][-1]) + 1
    last_col = int(np.nonzero(non_default.any(axis=0))[0][-1]) + 1
    return arr[: min(last_row + TRIM_MARGIN, arr.shape[0]),
               : min(last_col + TRIM_MARGIN, arr.shape[1])]


def load_sheet_features(
    file_path: str,
    sheet_name: str,
    max_rows: int = DEFAULT_MAX_ROWS,
    max_cols: int = DEFAULT_MAX_COLS,
) -> np.ndarray:
    """Convenience: open a workbook and featurize one sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path)
    try:
        return featurize_sheet(wb[sheet_name], max_rows=max_rows, max_cols=max_cols)
    finally:
        wb.close()

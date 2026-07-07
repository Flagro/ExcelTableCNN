"""Detection targets for spreadsheets.

Box convention (used everywhere in this package): ``[x_min, y_min, x_max,
y_max]`` in cell units, zero-indexed, **half-open** — ``x_max``/``y_max`` are
one past the last column/row the table occupies. ``"A1:C3"`` maps to
``[0, 0, 3, 3]``; a single cell ``"A1:A1"`` maps to ``[0, 0, 1, 1]``, so every
valid table has strictly positive width and height (torchvision rejects
degenerate boxes).
"""

import logging
from typing import Dict, List, Sequence

import torch
from torch.utils.data import Dataset
from openpyxl.utils.cell import get_column_letter, range_boundaries

logger = logging.getLogger(__name__)


def parse_table_range(table_range: str) -> List[float]:
    """Excel range string -> half-open [x_min, y_min, x_max, y_max] box."""
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    if None in (min_col, min_row, max_col, max_row):
        raise ValueError(f"Range {table_range!r} is not a bounded cell range")
    return [float(min_col - 1), float(min_row - 1), float(max_col), float(max_row)]


def box_to_range(box: Sequence[float]) -> str:
    """Half-open box -> Excel range string (inverse of parse_table_range)."""
    x_min, y_min, x_max, y_max = (int(round(float(v))) for v in box)
    x_min, y_min = max(x_min, 0), max(y_min, 0)
    x_max = max(x_max, x_min + 1)
    y_max = max(y_max, y_min + 1)
    return f"{get_column_letter(x_min + 1)}{y_min + 1}:{get_column_letter(x_max)}{y_max}"


def validate_boxes(boxes: torch.Tensor, height: int, width: int) -> torch.Tensor:
    """Clamp boxes to the tensor bounds and drop degenerate ones."""
    if boxes.numel() == 0:
        return boxes.reshape(0, 4)
    boxes = boxes.clone()
    boxes[:, 0::2] = boxes[:, 0::2].clamp(0, width)
    boxes[:, 1::2] = boxes[:, 1::2].clamp(0, height)
    keep = (boxes[:, 2] > boxes[:, 0]) & (boxes[:, 3] > boxes[:, 1])
    return boxes[keep]


class SpreadsheetDataset(Dataset):
    """Dataset over featurized sheet samples.

    Each sample is a dict with keys ``tensor`` (C×H×W float32), ``boxes``
    (N×4 float32, half-open convention), ``file_path`` and ``sheet_name``.
    Samples whose boxes are all invalid after clamping are skipped (recorded
    in ``self.skipped``), not crashed on.
    """

    def __init__(self, samples: Sequence[Dict]):
        self.samples: List[Dict] = []
        self.skipped: List[Dict] = []
        for sample in samples:
            tensor = sample["tensor"]
            _, height, width = tensor.shape
            boxes = validate_boxes(
                torch.as_tensor(sample["boxes"], dtype=torch.float32), height, width
            )
            if len(boxes) == 0:
                self.skipped.append(sample)
                continue
            self.samples.append({**sample, "boxes": boxes})
        if self.skipped:
            logger.warning(
                "Skipped %d sheet(s) with no valid table boxes (e.g. %s!%s)",
                len(self.skipped),
                self.skipped[0].get("file_path"),
                self.skipped[0].get("sheet_name"),
            )

    @property
    def num_cell_features(self) -> int:
        return self.samples[0]["tensor"].shape[0] if self.samples else 0

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, idx: int):
        sample = self.samples[idx]
        target = {
            "boxes": sample["boxes"],
            "labels": torch.ones((len(sample["boxes"]),), dtype=torch.int64),
        }
        return sample["tensor"], target


def collate_fn(batch):
    """Detection collate: keep variable-size tensors as tuples."""
    return tuple(zip(*batch))

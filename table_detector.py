import pandas as pd
from openpyxl.utils import coordinate_to_tuple, get_column_letter

from train_test_composer import get_train_data, get_backup_train_data, get_test_data, balance_train_data
from ml_classifier import predict_header_coordinates


def get_header_rectangle(header_coordinates, filter_rows=True, min_width=3):
    coordinates = set(map(tuple, map(coordinate_to_tuple, header_coordinates)))
    coloring_dict = dict()

    def coordinate_dfs_coloring(cur_coordinate, cur_color):
        coloring_dict[cur_coordinate] = cur_color
        relative_adjacent_shift = [(-1, -1), (-1, 0), (-1, 1),
                                   (0, -1), (0, 0), (0, 1),
                                   (1, -1), (1, 0), (1, 1)]
        for shift in relative_adjacent_shift:
            new_coordinate = (cur_coordinate[0] + shift[0], cur_coordinate[1] + shift[1])
            if new_coordinate in coordinates and new_coordinate not in coloring_dict:
                coordinate_dfs_coloring(new_coordinate, cur_color)

    color = 0
    for coordinate in coordinates:
        if coordinate not in coloring_dict:
            coordinate_dfs_coloring(coordinate, color)
            color += 1

    header_ranges = []
    for i in range(color):
        cur_header_coordinates = [coordinate for coordinate in coordinates if coloring_dict[coordinate] == i]

        if filter_rows:
            row_distribution = dict()
            for cur_row, cur_col in cur_header_coordinates:
                if cur_row not in row_distribution:
                    row_distribution[cur_row] = 0
                row_distribution[cur_row] += 1

            max_cols = max(row_distribution.values())
            rows_to_delete = [key for key, val in row_distribution.items() if val < max_cols / 2]
            cur_header_coordinates = [coordinate for coordinate in cur_header_coordinates
                                      if coordinate[0] not in rows_to_delete]

        min_row = min(list(map(lambda x: x[0], cur_header_coordinates)))
        max_row = max(list(map(lambda x: x[0], cur_header_coordinates)))
        min_col = min(list(map(lambda x: x[1], cur_header_coordinates)))
        max_col = max(list(map(lambda x: x[1], cur_header_coordinates)))

        first_point = get_column_letter(min_col) + str(min_row)
        second_point = get_column_letter(max_col) + str(max_row)
        if min_width is not None and max_col - min_col + 1 < min_width and color >= 2:
            continue

        header_ranges.append(f"{first_point}:{second_point}")

    return header_ranges


def get_table_body(header_range, other_headers, test_df, allowed_empty_rows=2):
    first_point, second_point = header_range.split(":")
    min_row, min_col = coordinate_to_tuple(first_point)
    max_row, max_col = coordinate_to_tuple(second_point)

    start_row = max_row + 1

    df = test_df.copy()

    df["is_not_empty"] = 1 - df["is_empty"]

    df["is_not_blank"] = df[["is_not_empty", "is_merged",
                             "left_border", "right_border", "top_border", "bottom_border"]].max(axis=1)

    df["row"] = df["coordinate"].apply(lambda x: coordinate_to_tuple(x)[0])
    df["column"] = df["coordinate"].apply(lambda x: coordinate_to_tuple(x)[1])

    df = df[["row", "column", "is_not_blank"]]

    stop_row = df["row"].max() + 1
    for other_header in other_headers:
        fp, sp = other_header.split(":")
        min_r, min_c = coordinate_to_tuple(fp)
        max_r, max_c = coordinate_to_tuple(sp)
        if min_r >= start_row and (min_col <= min_c <= max_col or min_col <= max_c <= max_col):
            if min_r < stop_row:
                stop_row = min_r

    filtered_df = df[(df["row"] >= start_row) & (df["row"] < stop_row) &
                     (df["column"] >= min_col) & (df["column"] <= max_col)]

    grouped_df = filtered_df.groupby("row").agg({"is_not_blank": max}).reset_index()

    result_df = pd.concat([grouped_df["row"], grouped_df["is_not_blank"]] +
                          [grouped_df["is_not_blank"].shift(i + 1) for i in range(1, allowed_empty_rows + 1)], axis=1)
    result_df = result_df.fillna(1)
    shift_columns = [f"is_not_blank_{i}" for i in range(1, allowed_empty_rows + 1)]
    result_df.columns = ["row", "is_not_blank"] + shift_columns
    result_df["allowed"] = result_df[shift_columns].max(axis=1)

    result_df["viable"] = (result_df["is_not_blank"] + result_df["allowed"] == 2)
    result_df["next_is_viable"] = result_df["viable"].shift(-1)
    result_df = result_df.fillna(False)

    result_df["result"] = (result_df["viable"] & ~result_df["next_is_viable"])

    result_row = result_df[result_df["result"]].reset_index(drop=True).loc[0]["row"]

    return "{}:{}".format(get_column_letter(min_col) + str(start_row),
                          get_column_letter(max_col) + str(result_row))


def get_tables(file_path, sheet_name, train_data=r".\\train\\train.pkl", mode="ml", retries=5):
    if train_data.endswith(".pkl"):
        train_df = get_backup_train_data(train_data)
    else:
        train_df = get_train_data(train_data)

    balanced_train_df = balance_train_data(train_df)
    # balanced_train_df = train_df

    test_df = get_test_data(file_path, sheet_name)

    header_ranges = []
    header_points = []
    for _ in range(retries):
        if mode == "ml":
            cur_header_points = predict_header_coordinates(balanced_train_df, test_df)
        else:
            cur_header_points = []
        cur_header_ranges = get_header_rectangle(cur_header_points)
        if not header_ranges or len(cur_header_ranges) < len(header_ranges):
            header_ranges = cur_header_ranges
            header_points = cur_header_points
        # print(header_ranges)

    print(header_points)
    body_ranges = []
    for header_range in header_ranges:
        body_range = get_table_body(header_range, header_ranges, test_df)
        body_ranges.append(body_range)

    result = [{"header": header_range, "body": body_range}
              for header_range, body_range in zip(header_ranges, body_ranges)]

    return result

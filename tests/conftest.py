"""Shared fixtures: a crafted .xlsx workbook with known cell features, and a
synthetic tensor sample factory for model tests that don't need Excel I/O."""

import datetime

import pytest
import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from excel_table_cnn.data.features import FEATURE_NAMES

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL = PatternFill(fill_type="solid", fgColor="FFFF00")

# Ground-truth table range of the toy workbook (headers B2:D2 + data B3:D6).
TOY_TABLE_RANGE = "B2:D6"
TOY_SHEET = "Data"


def feature_index(name: str) -> int:
    return FEATURE_NAMES.index(name)


@pytest.fixture()
def toy_workbook_path(tmp_path):
    """A workbook exercising every feature channel at known coordinates."""
    wb = Workbook()
    ws = wb.active
    ws.title = TOY_SHEET

    # A 3-column table at B2:D6: bold, bordered, filled, centered headers.
    for col, header in zip("BCD", ["Name", "Qty", "Price"]):
        cell = ws[f"{col}2"]
        cell.value = header
        cell.font = Font(b=True)
        cell.border = THIN_BORDER
        cell.fill = FILL
        cell.alignment = Alignment(horizontal="center")
    for row in range(3, 7):
        ws[f"B{row}"] = f"item-{row}"
        ws[f"C{row}"] = row * 10
        ws[f"D{row}"] = row * 1.5

    # Merged region F2:G3, value on the anchor cell only.
    ws["F2"] = "merged title"
    ws.merge_cells("F2:G3")

    # Assorted single-feature cells.
    ws["F6"] = "italic"
    ws["F6"].font = Font(i=True)
    ws["B8"] = "=SUM(C3:C6)"  # formula
    ws["F8"] = "wrapped text cell"
    ws["F8"].alignment = Alignment(wrapText=True)
    ws["G8"] = "indented"
    ws["G8"].alignment = Alignment(indent=1)
    ws["F10"] = "left"
    ws["F10"].alignment = Alignment(horizontal="left")
    ws["G10"] = "right"
    ws["G10"].alignment = Alignment(horizontal="right")

    # v2 feature cells: formats, colors, merge directions, text stats.
    ws["B12"] = 0.125
    ws["B12"].number_format = "0.00%"  # numeric (percent) format
    ws["C12"] = 3.14  # decimal point in value text
    ws["D12"] = datetime.date(2001, 3, 14)
    ws["D12"].number_format = "yyyy-mm-dd"
    ws["E12"] = datetime.time(9, 30)
    ws["E12"].number_format = "hh:mm"
    ws["F12"] = "red text"
    ws["F12"].font = Font(color="FFFF0000")
    ws["G12"] = "12.5%"  # percent symbol in value text
    ws["B14"] = "wide merge"
    ws.merge_cells("B14:D14")
    ws["F14"] = "tall merge"
    ws.merge_cells("F14:F16")

    path = tmp_path / "toy.xlsx"
    wb.save(path)
    wb.close()
    return str(path)


@pytest.fixture()
def toy_xls_path(tmp_path):
    """Legacy .xls twin of the toy workbook (written with xlwt, read with
    xlrd — no LibreOffice involved). Mirrors ``toy_workbook_path`` cell for
    cell, except the formula cell: the formula channel is undetectable in
    .xls and stays 0."""
    import xlwt

    wb = xlwt.Workbook()
    ws = wb.add_sheet(TOY_SHEET)

    header_style = xlwt.easyxf(
        "font: bold on; "
        "borders: left thin, right thin, top thin, bottom thin; "
        "pattern: pattern solid, fore_colour yellow; "
        "align: horiz center"
    )
    for col, header in zip((1, 2, 3), ["Name", "Qty", "Price"]):
        ws.write(1, col, header, header_style)  # row 2 in Excel terms
    for row in range(2, 6):  # Excel rows 3..6
        ws.write(row, 1, f"item-{row + 1}")
        ws.write(row, 2, (row + 1) * 10)
        ws.write(row, 3, (row + 1) * 1.5)

    ws.write_merge(1, 2, 5, 6, "merged title")  # F2:G3
    ws.write(5, 5, "italic", xlwt.easyxf("font: italic on"))  # F6
    ws.write(7, 5, "wrapped text cell", xlwt.easyxf("align: wrap on"))  # F8
    ws.write(7, 6, "indented", xlwt.easyxf("align: indent 1"))  # G8
    ws.write(9, 5, "left", xlwt.easyxf("align: horiz left"))  # F10
    ws.write(9, 6, "right", xlwt.easyxf("align: horiz right"))  # G10

    # v2 feature cells, mirroring the xlsx fixture.
    ws.write(11, 1, 0.125, xlwt.easyxf(num_format_str="0.00%"))  # B12
    ws.write(11, 2, 3.14)  # C12
    ws.write(11, 3, datetime.date(2001, 3, 14),
             xlwt.easyxf(num_format_str="YYYY-MM-DD"))  # D12
    ws.write(11, 4, datetime.time(9, 30), xlwt.easyxf(num_format_str="hh:mm"))  # E12
    ws.write(11, 5, "red text", xlwt.easyxf("font: colour red"))  # F12
    ws.write(11, 6, "12.5%")  # G12
    ws.write_merge(13, 13, 1, 3, "wide merge")  # B14:D14
    ws.write_merge(13, 15, 5, 5, "tall merge")  # F14:F16

    path = tmp_path / "toy.xls"
    wb.save(str(path))
    return str(path)


def make_synthetic_sample(height=30, width=15, box=(3, 5, 11, 21)):
    """A sheet tensor with one clearly featurized table at ``box`` (half-open
    [x_min, y_min, x_max, y_max]). Outside the box: empty cells."""
    tensor = torch.zeros((len(FEATURE_NAMES), height, width), dtype=torch.float32)
    tensor[feature_index("is_empty")] = 1.0

    x_min, y_min, x_max, y_max = box
    tensor[feature_index("is_empty"), y_min:y_max, x_min:x_max] = 0.0
    tensor[feature_index("is_string"), y_min:y_max, x_min:x_max] = 1.0
    # Header row: bold + filled + bordered.
    tensor[feature_index("is_bold"), y_min, x_min:x_max] = 1.0
    tensor[feature_index("is_filled"), y_min, x_min:x_max] = 1.0
    tensor[feature_index("top_border"), y_min, x_min:x_max] = 1.0
    tensor[feature_index("bottom_border"), y_max - 1, x_min:x_max] = 1.0

    return {
        "tensor": tensor,
        "boxes": torch.tensor([list(map(float, box))], dtype=torch.float32),
        "file_path": "<synthetic>",
        "sheet_name": "synthetic",
        "feature_names": list(FEATURE_NAMES),
    }

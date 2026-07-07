import numpy as np
import openpyxl
import pytest
from openpyxl import Workbook

from excel_table_cnn.data.features import (
    FEATURE_NAMES,
    NUM_FEATURES,
    default_cell_vector,
    featurize_sheet,
    load_sheet_features,
)
from .conftest import TOY_SHEET, feature_index


def cell(arr, ref):
    """Feature vector of an Excel-style cell reference in the array."""
    from openpyxl.utils.cell import coordinate_to_tuple

    row, col = coordinate_to_tuple(ref)
    return arr[row - 1, col - 1]


@pytest.fixture()
def toy_features(toy_workbook_path):
    return load_sheet_features(toy_workbook_path, TOY_SHEET)


def test_shape_and_dtype(toy_features):
    assert toy_features.ndim == 3
    assert toy_features.shape[2] == NUM_FEATURES
    assert toy_features.dtype == np.float32
    # Used range reaches G10 at least (content), so >= 10 rows and 7 cols.
    assert toy_features.shape[0] >= 10
    assert toy_features.shape[1] >= 7


def test_header_cell_features(toy_features):
    b2 = cell(toy_features, "B2")
    assert b2[feature_index("is_empty")] == 0
    assert b2[feature_index("is_string")] == 1
    assert b2[feature_index("is_bold")] == 1
    assert b2[feature_index("is_filled")] == 1
    for side in ("left_border", "right_border", "top_border", "bottom_border"):
        assert b2[feature_index(side)] == 1
    assert b2[feature_index("horizontal_alignment")] == 1
    assert b2[feature_index("center_horizontal_alignment")] == 1
    assert b2[feature_index("left_horizontal_alignment")] == 0
    assert b2[feature_index("formula")] == 0


def test_numeric_cell_is_not_string(toy_features):
    c3 = cell(toy_features, "C3")
    assert c3[feature_index("is_empty")] == 0
    assert c3[feature_index("is_string")] == 0


def test_untouched_cell_is_default(toy_features):
    e4 = cell(toy_features, "E4")
    assert (e4 == default_cell_vector()).all()


def test_borders_not_constant_true(toy_features):
    """Regression for the openpyxl border bug: default cells must NOT report
    borders (``cell.border.left`` is always a Side object; only its style
    indicates a drawn border)."""
    e4 = cell(toy_features, "E4")
    for side in ("left_border", "right_border", "top_border", "bottom_border"):
        assert e4[feature_index(side)] == 0
    border_channel = toy_features[:, :, feature_index("left_border")]
    assert border_channel.sum() < border_channel.size  # not constant 1


def test_merged_cells(toy_features):
    assert cell(toy_features, "F2")[feature_index("is_merged")] == 1  # anchor
    assert cell(toy_features, "G3")[feature_index("is_merged")] == 1  # covered
    assert cell(toy_features, "G3")[feature_index("is_empty")] == 1  # no own value
    assert cell(toy_features, "B2")[feature_index("is_merged")] == 0


def test_assorted_features(toy_features):
    assert cell(toy_features, "F6")[feature_index("is_italic")] == 1
    assert cell(toy_features, "B8")[feature_index("formula")] == 1
    assert cell(toy_features, "B8")[feature_index("is_string")] == 0
    assert cell(toy_features, "F8")[feature_index("wrapped_text")] == 1
    assert cell(toy_features, "G8")[feature_index("indent")] == 1
    assert cell(toy_features, "F10")[feature_index("left_horizontal_alignment")] == 1
    assert cell(toy_features, "G10")[feature_index("right_horizontal_alignment")] == 1


def test_trailing_default_region_is_trimmed():
    wb = Workbook()
    ws = wb.active
    for ref in ("A1", "B2", "C3"):
        ws[ref] = "x"
    ws["J20"] = None  # creates a default cell far away, inflating the used range
    assert ws.max_row == 20 and ws.max_column == 10

    arr = featurize_sheet(ws)
    # Content ends at C3; TRIM_MARGIN=2 keeps two extra rows/cols.
    assert arr.shape[:2] == (5, 5)


def test_used_range_cap_warns_and_clips(tmp_path, caplog):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["A400"] = "y"
    path = tmp_path / "cap.xlsx"
    wb.save(path)

    wb2 = openpyxl.load_workbook(str(path))
    with caplog.at_level("WARNING"):
        arr = featurize_sheet(wb2[ws.title], max_rows=100, max_cols=50)
    assert arr.shape[0] <= 100
    assert any("exceeds cap" in record.message for record in caplog.records)


def test_feature_names_stable():
    """Channel order is a compatibility contract for cached tensors and
    checkpoints — changing it must be a deliberate, version-bumped act."""
    assert NUM_FEATURES == 30
    assert FEATURE_NAMES[0] == "is_empty"
    assert FEATURE_NAMES[17] == "string_length"  # v1 prefix preserved
    assert len(set(FEATURE_NAMES)) == NUM_FEATURES


def test_value_string_statistics(toy_features):
    b2 = cell(toy_features, "B2")  # "Name"
    assert 0 < b2[feature_index("string_length")] < 1
    assert b2[feature_index("letter_ratio")] == 1.0
    assert b2[feature_index("digit_ratio")] == 0.0

    c12 = cell(toy_features, "C12")  # 3.14
    assert c12[feature_index("has_decimal_point")] == 1
    assert c12[feature_index("digit_ratio")] == 3 / 4

    g12 = cell(toy_features, "G12")  # "12.5%"
    assert g12[feature_index("has_percent_symbol")] == 1
    assert g12[feature_index("has_decimal_point")] == 1


def test_number_format_templates(toy_features):
    b12 = cell(toy_features, "B12")  # 0.00%
    assert b12[feature_index("numeric_format")] == 1
    assert b12[feature_index("date_format")] == 0
    assert 0 < b12[feature_index("format_length")] < 1

    d12 = cell(toy_features, "D12")  # yyyy-mm-dd
    assert d12[feature_index("date_format")] == 1
    assert d12[feature_index("time_format")] == 0

    e12 = cell(toy_features, "E12")  # hh:mm
    assert e12[feature_index("time_format")] == 1
    assert e12[feature_index("date_format")] == 0

    b2 = cell(toy_features, "B2")  # General
    assert b2[feature_index("numeric_format")] == 0
    assert b2[feature_index("format_length")] == 0


def test_merge_direction_channels(toy_features):
    b14 = cell(toy_features, "B14")  # B14:D14 — wide only
    assert b14[feature_index("merged_horizontal")] == 1
    assert b14[feature_index("merged_vertical")] == 0

    f14 = cell(toy_features, "F14")  # F14:F16 — tall only
    assert f14[feature_index("merged_horizontal")] == 0
    assert f14[feature_index("merged_vertical")] == 1

    f2 = cell(toy_features, "F2")  # F2:G3 — both
    assert f2[feature_index("merged_horizontal")] == 1
    assert f2[feature_index("merged_vertical")] == 1


def test_color_channels(toy_features):
    b2 = cell(toy_features, "B2")  # yellow fill, default font color
    assert b2[feature_index("non_default_fill_color")] == 1
    assert b2[feature_index("non_default_font_color")] == 0

    f12 = cell(toy_features, "F12")  # red font, no fill
    assert f12[feature_index("non_default_font_color")] == 1
    assert f12[feature_index("non_default_fill_color")] == 0

    e4 = cell(toy_features, "E4")  # untouched
    assert e4[feature_index("non_default_fill_color")] == 0
    assert e4[feature_index("non_default_font_color")] == 0


def test_date_format_not_fooled_by_color_section():
    """'[Red]' contains 'd' — the format classifier must strip sections."""
    from excel_table_cnn.data.features import format_features

    numeric, date, time, _ = format_features("#,##0_);[Red](#,##0)")
    assert numeric == 1 and date == 0 and time == 0
